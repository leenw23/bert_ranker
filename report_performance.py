import json
import os
import pickle
from pprint import pprint
from string import ascii_uppercase

# pip install uncertainty-calibration
import calibration as cal
import numpy as np
from sklearn.metrics import brier_score_loss, confusion_matrix, f1_score
from sklearn.metrics import precision_recall_fscore_support as score_metric

from utils import brier_multi, recall_x_at_k


def main_script(dirname, num_candidates):
    dirlist = os.listdir(dirname)
    dirlist = sorted([os.path.join(dirname, el) for el in dirlist])
    result = {}

    for exp_dir in dirlist:
        if exp_dir not in result:
            result[exp_dir] = {}

        flist = os.listdir(exp_dir)
        flist = [os.path.join(exp_dir, fname) for fname in flist]
        assert all([".json" in el for el in flist])
        flist = sorted(flist)

        for fname in flist:
            if str(num_candidates) in fname:
                assert "candi" in fname
                recall, ece, brier = main(fname, num_candidates)
                done = False
                for modelname in ["select", "temp", "mcdrop", "ensemble"]:
                    if modelname in fname:
                        done = True
                        break
                if not done:
                    raise ValueError
                assert "test" in fname or "dev" in fname
                if "dev" in fname:
                    print("$$ DEV  $$")
                    print(modelname)
                    print(recall, ece, brier)
                    print("\n")
                    continue

            result[exp_dir][modelname] = {"recall": recall, "ece": ece, "brier": brier}

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active

    ws["B1"] = "Select"
    ws["E1"] = "Temp"
    ws["H1"] = "MCDrop"
    ws["K1"] = "Ensemble"

    count = 3

    for exp_name, exp_result in result.items():
        ws["A" + str(count)] = exp_name.split("/")[-1]
        for model_name, model_result in exp_result.items():
            assert model_name in ["select", "temp", "mcdrop", "ensemble"]
            model_index = ["select", "temp", "mcdrop", "ensemble"].index(model_name)
            column_index = [1, 4, 7, 10][model_index]
            ws[ascii_uppercase[column_index] + str(count)] = model_result["recall"]
            ws[ascii_uppercase[column_index + 1] + str(count)] = model_result["ece"]
            ws[ascii_uppercase[column_index + 2] + str(count)] = model_result["brier"]
        count += 1
    wb.save("result/performance/performance_{}.xlsx".format(num_candidates))

    with open("result/dump_result_{}.json".format(num_candidates), "w") as f:
        json.dump(result, f)


def softmax_np(logits):
    exp_logits = np.exp(logits)
    probs = exp_logits / np.sum(exp_logits)

    return [float(el) for el in probs]


def main(fname, num_candidates):
    assert ".json" in fname

    with open(fname, "r") as f:
        prediction_data = [json.loads(el) for el in f.readlines() if el.strip() != ""]

    r10 = run_origianl_recall(prediction_data, num_candidates)
    
    calibration_error = cal.get_ece(
        [softmax_np(l["pred"]) for l in prediction_data],
        [0 for _ in range(len(prediction_data))],
    )
    brier_score = brier_multi(
        [[1] + [0 for _ in range(num_candidates-1)] for __ in range(len(prediction_data))],
        [softmax_np(l["pred"]) for l in prediction_data],
    )

    # print(fname)
    # print("R10@1: {}".format(r10))
    # print("ECE-R10@1: {}".format(calibration_error))
    # print("Brier: {}".format(brier_score))
    
    return r10, calibration_error, brier_score


def run_origianl_recall(
    prediction_list,
    x: int,
    ):
    """.

    Args:
        prediction_list (List[Dict[str,Union[List[float], bool]]]): {
                    "pred": [list of unnormalized score],
                    "uncertainty": [list of unnormalized uncertainty],
                    "is_uw": bool,
                }
    """
    Recall_list = []
    for item in prediction_list:
        uncertainty = item["uncertainty"][:x]
        prediction_outcome = item["pred"][:x]
        Recall_list.append(recall_x_at_k(prediction_outcome, x, 1, 0))
    return sum(Recall_list) / len(Recall_list)


if __name__ == "__main__":
    num_candidates = 10
    dirname = "./result/"
    main_script(dirname, num_candidates)
